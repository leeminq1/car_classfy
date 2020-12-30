from pymongo import MongoClient
import jwt
import datetime
import hashlib
from flask import Flask, render_template, jsonify, request, redirect, url_for
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta

app = Flask(__name__)
SECRET_KEY = 'SPARTA'

# client = MongoClient('localhost', 27017)
client = MongoClient('mongodb://test:test@localhost',27017)
db = client.dbsparta


from openpyxl import load_workbook

# data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook("car_list_recent.xlsx", data_only=True)
# 시트 이름으로 불러오기
load_ws = load_wb['Sheet1']


allList = []
for row in load_ws.iter_rows(min_row=2, max_row=55, min_col=1, max_col=13):
    a = []
    for cell in row:
        a.append(cell.value)
    allList.append(a)
for item in allList:
    doc={"name":item[2],
         "number":item[0],
         "location":item[3],
         "class":item[7],
         "comment":None,
         "location_number":item[4],
         "car_admin":item[5],
         "car_region":item[6],
         "car_temp_select":item[8],
         "car_temp_number":str(item[9]),
         "temp_number_start":str(item[10]),
         "temp_number_end": str(item[11]),
         "specific_caruse": str(item[12])
        }
    print(doc)
    db.cardb.insert_one(doc)